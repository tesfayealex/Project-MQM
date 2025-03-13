from rest_framework import viewsets, permissions, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response as DRFResponse
from django.db.models import Count, Avg, Q, F, Sum, FloatField, Case, When, Value, ExpressionWrapper
from django.db.models.functions import Cast
from django.utils import timezone
from datetime import timedelta
from django.contrib.auth.models import User, Group
from collections import Counter
from .models import Survey, Question, Response, Answer, SurveyToken, WordCluster, ResponseWord, SurveyAnalysisSummary, CustomWordCluster, Template
from .serializers import (
    SurveySerializer, 
    SurveyDetailSerializer,
    QuestionSerializer, 
    ResponseSerializer, 
    AnswerSerializer,
    SurveyTokenSerializer,
    WordClusterSerializer,
    ResponseWordSerializer,
    SurveyAnalysisSummarySerializer,
    CustomWordClusterSerializer,
    TemplateSerializer,
    TemplateDetailSerializer,
    SurveyWithTemplateSerializer
)
from django.http import HttpResponse
import qrcode
from io import BytesIO
from django.db import transaction
import logging
from .utils import (
    TextAnalyzer, cluster_responses, calculate_stats_from_scores, 
    calculate_satisfaction_score, process_text, process_survey_and_assign_clusters, assign_clusters_to_words,
    analyze_response_clusters, get_survey_sentence_sentiment_analysis, analyze_sentences, process_sentence
)
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
import numpy as np
from statistics import stdev, mean, median
from .direct_process_all_responses import direct_process_all_responses
from openpyxl import Workbook
from rest_framework.renderers import BaseRenderer

# Configure logger
logger = logging.getLogger(__name__)

class BinaryFileRenderer(BaseRenderer):
    """
    Renderer for binary files like Excel
    """
    media_type = 'application/octet-stream'
    format = 'binary'
    charset = None
    render_style = 'binary'

    def render(self, data, media_type=None, renderer_context=None):
        # Return the raw binary data as-is
        return data

class IsCreatorOrReadOnly(permissions.BasePermission):
    def has_permission(self, request, view):
        # Must be authenticated
        if not request.user.is_authenticated:
            return False
            
        # Admin and Organizer have full access
        if request.user.groups.filter(name__in=['Admin', 'Organizer']).exists():
            return True
            
        # Moderator can create surveys and access their own
        if request.user.groups.filter(name='Moderator').exists():
            return True
            
        # For GET requests, allow if survey is active (for participants)
        if request.method in permissions.SAFE_METHODS:
            return True
            
        return False

    def has_object_permission(self, request, view, obj):
        # Admin and Organizer have full access
        if request.user.groups.filter(name__in=['Admin', 'Organizer']).exists():
            return True
            
        # Moderator can only view and edit (but not delete) their own surveys
        if request.user.groups.filter(name='Moderator').exists():
            if request.method == 'DELETE':
                return False
            return obj.created_by == request.user
            
        # For GET requests, allow if survey is active (for participants)
        if request.method in permissions.SAFE_METHODS:
            return obj.is_active
            
        return False


class SurveyViewSet(viewsets.ModelViewSet):
    queryset = Survey.objects.all()
    permission_classes = [permissions.IsAuthenticated, IsCreatorOrReadOnly]
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return SurveyWithTemplateSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return SurveyDetailSerializer
        return SurveySerializer
    
    def get_queryset(self):
        queryset = Survey.objects.all()
        
        # Admin and Organizer can see all surveys
        if self.request.user.groups.filter(name__in=['Admin', 'Organizer']).exists():
            pass
        # Moderator can only see their own surveys
        elif self.request.user.groups.filter(name='Moderator').exists():
            queryset = queryset.filter(created_by=self.request.user)
        # Others (participants) can only see active surveys
        else:
            queryset = queryset.filter(is_active=True)
        
        # Filter by creator if requested
        created_by = self.request.query_params.get('created_by', None)
        if created_by and created_by == 'me':
            queryset = queryset.filter(created_by=self.request.user)
            
        # Filter by language if requested
        language = self.request.query_params.get('language', None)
        if language:
            queryset = queryset.filter(languages__contains=[language])
            
        # Search by title or description
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) | 
                Q(description__icontains=search)
            )
            
        return queryset.order_by('-created_at')
    
    @action(detail=True, methods=['get'])
    def stats(self, request, pk=None):
        survey = self.get_object()
        
        # Get number of responses
        total_responses = Response.objects.filter(survey=survey).count()
        
        # Get responses by language
        responses_by_language = Response.objects.filter(survey=survey).values('language').annotate(count=Count('id'))
        
        # Calculate average NPS score
        nps_avg = Answer.objects.filter(
            question__survey=survey,
            question__type='nps',
            nps_rating__isnull=False
        ).aggregate(avg_score=Avg('nps_rating'))['avg_score'] or 0
        
        # Calculate NPS categories
        nps_answers = Answer.objects.filter(
            question__survey=survey,
            question__type='nps',
            nps_rating__isnull=False
        )
        
        total_nps = nps_answers.count()
        promoters = nps_answers.filter(nps_rating__gte=9).count()
        detractors = nps_answers.filter(nps_rating__lte=6).count()
        
        nps_score = 0
        if total_nps > 0:
            nps_score = ((promoters / total_nps) - (detractors / total_nps)) * 100
        
        # Calculate completion rate
        completion_rate = self.calculate_completion_rate(survey)
        
        return DRFResponse({
            'total_responses': total_responses,
            'responses_by_language': list(responses_by_language),
            'nps_average': round(nps_avg, 1),
            'nps_score': round(nps_score, 1),
            'completion_rate': round(completion_rate, 1),
            'promoters': promoters,
            'detractors': detractors,
            'passives': total_nps - promoters - detractors
        })
    
    @action(detail=False, methods=['get'], permission_classes=[permissions.AllowAny], url_path='public')
    def public(self, request):
        token = request.query_params.get('token')
        
        if not token:
            return DRFResponse({'detail': 'Token is required'}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            # Find survey by token - looking at both legacy token field and SurveyToken model
            try:
                # First try to find a survey using the SurveyToken model
                survey_token = SurveyToken.objects.get(token=token)
                survey = survey_token.survey
            except SurveyToken.DoesNotExist:
                # If not found, try the legacy token field
                survey = Survey.objects.get(token=token)
            
            serializer = SurveyDetailSerializer(survey)
            survey_data = serializer.data
            
            # Check if survey is active and not expired
            if not survey.is_active:
                return DRFResponse({
                    'detail': 'This survey is no longer active',
                    'survey': survey_data,
                    'status': 'inactive'
                }, status=status.HTTP_400_BAD_REQUEST)
                
            # Check if survey has expired
            if survey.expiry_date and survey.expiry_date < timezone.now():
                return DRFResponse({
                    'detail': 'This survey has expired',
                    'survey': survey_data,
                    'status': 'expired'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            return DRFResponse(survey_data)
            
        except (Survey.DoesNotExist, SurveyToken.DoesNotExist):
            return DRFResponse({'detail': 'Survey not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def qr_code_data(self, request, pk=None):
        survey = self.get_object()
        
        # Check if survey has any tokens
        token_objects = survey.tokens.all()
        if not token_objects.exists() and not survey.token:
            return DRFResponse(
                {"error": "Survey does not have any tokens"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        base_url = request.build_absolute_uri('/')[:-1]  # Remove trailing slash
        
        # Get all token data including legacy token if present
        token_data = []
        
        # Add SurveyToken objects
        for token_obj in token_objects:
            token_data.append({
                'id': token_obj.id,
                'token': token_obj.token,
                'description': token_obj.description,
                'survey_url': f"{base_url}/survey/{token_obj.token}",
                'qr_code_url': f"{base_url}/api/surveys/surveys/token/{token_obj.token}/qr_code/"
            })
        
        # If there's a legacy token, include it too
        if survey.token and not token_objects.filter(token=survey.token).exists():
            token_data.append({
                'id': None,
                'token': survey.token,
                'description': 'Legacy Token',
                'survey_url': f"{base_url}/survey/{survey.token}",
                'qr_code_url': f"{base_url}/api/surveys/surveys/token/{survey.token}/qr_code/"
            })
        
        # If no tokens exist, return error
        if not token_data:
            return DRFResponse(
                {"error": "Survey does not have any tokens"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Return all token data
        return DRFResponse({
            'tokens': token_data,
            'primary_token': survey.primary_token
        })

    @action(detail=False, methods=['get'], permission_classes=[permissions.AllowAny], url_path='token/(?P<token>[^/.]+)/qr_code')
    def token_qr_code(self, request, token=None):
        """Generate QR code for a specific token"""
        if not token:
            return DRFResponse({'detail': 'Token is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Find survey by token - looking at both legacy token field and SurveyToken model
            try:
                # First try to find a survey using the SurveyToken model
                survey_token = SurveyToken.objects.get(token=token)
                survey = survey_token.survey
            except SurveyToken.DoesNotExist:
                # If not found, try the legacy token field
                survey = Survey.objects.get(token=token)
            
            # Generate the URL for the public survey
            base_url = request.build_absolute_uri('/').rstrip('/')
            survey_url = f"{base_url}/api/surveys/public?token={token}"
            
            # Generate QR code
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            
            qr.add_data(survey_url)
            qr.make(fit=True)
            
            img = qr.make_image(fill_color="black", back_color="white")
            
            # Save QR code to BytesIO object
            buffer = BytesIO()
            img.save(buffer, format="PNG")
            buffer.seek(0)
            
            # Return the image
            return HttpResponse(buffer.getvalue(), content_type="image/png")
            
        except (Survey.DoesNotExist, SurveyToken.DoesNotExist):
            return DRFResponse({'detail': 'Survey not found'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['get'], permission_classes=[permissions.AllowAny])
    def qr_code(self, request, pk=None):
        """Legacy endpoint for backward compatibility - uses the primary token"""
        try:
            survey = self.get_object()
            
            # Get the primary token
            primary_token = survey.primary_token
            
            if not primary_token:
                return DRFResponse({'detail': 'This survey does not have a token for public access'}, status=status.HTTP_400_BAD_REQUEST)
                
            # Generate the URL for the public survey
            base_url = request.build_absolute_uri('/').rstrip('/')
            survey_url = f"{base_url}/api/surveys/public?token={primary_token}"
            
            # Generate QR code
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            
            qr.add_data(survey_url)
            qr.make(fit=True)
            
            img = qr.make_image(fill_color="black", back_color="white")
            
            # Save QR code to BytesIO object
            buffer = BytesIO()
            img.save(buffer, format="PNG")
            buffer.seek(0)
            
            # Return the image
            return HttpResponse(buffer.getvalue(), content_type="image/png")
            
        except Survey.DoesNotExist:
            return DRFResponse({'detail': 'Survey not found'}, status=status.HTTP_404_NOT_FOUND)

    # New endpoint for managing tokens
    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def add_token(self, request, pk=None):
        """Add a new token to the survey"""
        survey = self.get_object()
        
        # Validate token data
        serializer = SurveyTokenSerializer(data=request.data)
        if serializer.is_valid():
            # Check if token already exists
            token = serializer.validated_data['token']
            if SurveyToken.objects.filter(token=token).exists() or Survey.objects.filter(token=token).exclude(id=survey.id).exists():
                return DRFResponse({'detail': 'Token already exists'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Create the token
            serializer.save(survey=survey)
            return DRFResponse(serializer.data, status=status.HTTP_201_CREATED)
        
        return DRFResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['delete'], permission_classes=[permissions.IsAuthenticated], url_path='token/(?P<token_id>\d+)')
    def delete_token(self, request, pk=None, token_id=None):
        """Delete a token from the survey"""
        survey = self.get_object()
        
        try:
            token = SurveyToken.objects.get(id=token_id, survey=survey)
            
            # Check if this is the only token
            if survey.tokens.count() <= 1:
                return DRFResponse({'detail': 'Cannot delete the only token. Surveys must have at least one token.'}, 
                                  status=status.HTTP_400_BAD_REQUEST)
            
            token.delete()
            return DRFResponse(status=status.HTTP_204_NO_CONTENT)
        
        except SurveyToken.DoesNotExist:
            return DRFResponse({'detail': 'Token not found'}, status=status.HTTP_404_NOT_FOUND)

    def calculate_completion_rate(self, survey):
        total_starts = Response.objects.filter(survey=survey).count()
        if total_starts == 0:
            return 0
        
        # Count responses that have answers for all required questions
        required_questions = Question.objects.filter(survey=survey, is_required=True).count()
        
        if required_questions == 0:
            return 100
        
        completed = 0
        for response in Response.objects.filter(survey=survey):
            answer_count = Answer.objects.filter(
                response=response, 
                question__is_required=True
            ).count()
            
            if answer_count >= required_questions:
                completed += 1
        
        return (completed / total_starts) * 100

    def perform_destroy(self, instance):
        # Additional check for delete permission
        if self.request.user.groups.filter(name='Moderator').exists():
            raise permissions.PermissionDenied("Moderators cannot delete surveys")
        super().perform_destroy(instance)

    @action(detail=True, methods=['get'], renderer_classes=[BinaryFileRenderer])  # Use BinaryFileRenderer instead of empty list
    def export_responses(self, request, pk=None):
        """
        Export survey responses to Excel
        """
        try:
            print(f"\n==== EXPORT RESPONSES DEBUG ====")
            print(f"Request method: {request.method}")
            import pandas as pd
            import io
            from openpyxl.utils import get_column_letter
            from django.http import HttpResponse
            
            # Direct console output for debugging
            print(f"\n\n==== EXPORT RESPONSES DEBUG ====")
            print(f"Request method: {request.method}")
            print(f"Survey ID: {pk}")
            print(f"User: {request.user}")
            
            survey = Survey.objects.get(pk=pk)
            self.check_object_permissions(request, survey)
            
            # Get all responses for this survey
            responses = Response.objects.filter(survey=survey).prefetch_related(
                'answers__question'
            ).order_by('-created_at')
            
            print(f"Found {responses.count()} responses to export")
            
            if not responses.exists():
                print("No responses found - returning 404")
                return DRFResponse({'detail': 'No responses found'}, status=status.HTTP_404_NOT_FOUND)
            
            # Prepare data for Excel export
            data = []
            
            # First, get all unique questions to use as columns
            questions = Question.objects.filter(survey=survey).order_by('order')
            print(f"Found {questions.count()} questions")
            
            # Create headers
            headers = ['Response ID', 'Session ID', 'Date', 'Language']
            question_headers = []
            
            for question in questions:
                # Use question text in the primary survey language or the first available language
                for lang in survey.languages:
                    if lang in question.questions:
                        question_headers.append(question.questions[lang])
                        break
                else:
                    # Fallback if no matching language found
                    question_headers.append(f"Question {question.id}")
            
            headers.extend(question_headers)
            
            # Add data for each response
            for response in responses:
                row = [
                    response.id,
                    response.session_id,
                    response.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    response.language,
                ]
                
                # Create a dictionary of question_id -> answer for easy lookup
                answer_dict = {}
                for answer in response.answers.all():
                    if answer.question_id:
                        answer_dict[answer.question_id] = answer
                
                # Add answers in the order of questions
                for question in questions:
                    if question.id in answer_dict:
                        answer = answer_dict[question.id]
                        if answer.nps_rating is not None:
                            row.append(answer.nps_rating)
                        elif answer.text_answer:
                            row.append(answer.text_answer)
                        else:
                            row.append('')
                    else:
                        row.append('')
                
                data.append(row)
            
            # Create DataFrame and Excel file
            df = pd.DataFrame(data, columns=headers)
            
            # Create a buffer for the Excel file
            buffer = io.BytesIO()
            
            # Create Excel writer
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Responses')
                
                # Auto-adjust columns width
                worksheet = writer.sheets['Responses']
                for i, col in enumerate(df.columns):
                    # Find the maximum length of the column
                    max_length = max(
                        df[col].astype(str).map(len).max(),
                        len(str(col))
                    ) + 2  # Add a little extra space
                    
                    # Limit column width to avoid extremely wide columns
                    column_width = min(max_length, 50)
                    
                    # Set the column width
                    worksheet.column_dimensions[get_column_letter(i+1)].width = column_width
            
            # Set up the response with the file
            buffer.seek(0)
            
            print(f"Excel file created, buffer size: {len(buffer.getvalue())}")
            
            # IMPORTANT CHANGE: Return a direct HttpResponse, completely bypassing DRF
            # This avoids content negotiation issues
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="survey-responses-{pk}.xlsx"'
            
            print(f"Returning Excel file with headers: {dict(response.headers)}")
            return response  # Return HttpResponse directly, not DRFResponse
            
        except Survey.DoesNotExist:
            print(f"Survey {pk} not found")
            return DRFResponse({'detail': 'Survey not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            import traceback
            error_traceback = traceback.format_exc()
            print(f"Error exporting responses: {str(e)}")
            print(f"Error type: {type(e).__name__}")
            print(f"Error traceback: {error_traceback}")
            logger.error(f"Error exporting responses: {str(e)}", exc_info=True)
            return DRFResponse({'detail': f"Error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    permission_classes = [permissions.IsAuthenticated, IsCreatorOrReadOnly]

    def get_queryset(self):
        queryset = Question.objects.all()
        
        # Admin and Organizer can see all questions
        if self.request.user.groups.filter(name__in=['Admin', 'Organizer']).exists():
            pass
        # Moderator can only see questions from their surveys
        elif self.request.user.groups.filter(name='Moderator').exists():
            queryset = queryset.filter(survey__created_by=self.request.user)
        # Others can only see questions from active surveys
        else:
            queryset = queryset.filter(survey__is_active=True)
        
        # Filter by survey if requested
        survey_id = self.request.query_params.get('survey', None)
        if survey_id is not None:
            queryset = queryset.filter(survey_id=survey_id)
            
        # Filter by language if requested
        language = self.request.query_params.get('language', None)
        if language is not None:
            queryset = queryset.filter(language=language)
            
        return queryset.order_by('order')

    def perform_create(self, serializer):
        survey_id = self.request.data.get('survey')
        if survey_id:
            survey = Survey.objects.get(id=survey_id)
            # Check if user has permission to add questions
            if not self.request.user.groups.filter(name__in=['Admin', 'Organizer']).exists():
                if survey.created_by != self.request.user:
                    raise permissions.PermissionDenied("You don't have permission to add questions to this survey.")
        serializer.save()

    def perform_destroy(self, instance):
        # Moderators cannot delete questions
        if self.request.user.groups.filter(name='Moderator').exists():
            raise permissions.PermissionDenied("Moderators cannot delete questions")
        super().perform_destroy(instance)


class ResponseViewSet(viewsets.ModelViewSet):
    queryset = Response.objects.all()
    serializer_class = ResponseSerializer
    permission_classes = [permissions.AllowAny]  # Allow public submissions

    def get_queryset(self):
        queryset = Response.objects.all()
        
        # Only show responses for surveys created by the current user if they're not an admin
        if not self.request.user.is_staff and self.request.user.is_authenticated:
            queryset = queryset.filter(survey__created_by=self.request.user)
            
        survey_id = self.request.query_params.get('survey', None)
        if survey_id is not None:
            queryset = queryset.filter(survey_id=survey_id)
            
        language = self.request.query_params.get('language', None)
        if language is not None:
            queryset = queryset.filter(language=language)
            
        return queryset.order_by('-created_at')

    @action(detail=True, methods=['get'])
    def extracted_words(self, request, pk=None):
        """
        Retrieve all extracted words for a specific response.
        """
        response = self.get_object()
        from .models import ResponseWord
        
        # Get all extracted words for this response
        words = ResponseWord.objects.filter(response=response)
        
        # Create a serializable format
        result = []
        for word in words:
            # Get the custom cluster name if assigned
            assigned_cluster = word.assigned_cluster
            
            # If no directly assigned cluster, check if it belongs to any custom clusters
            if not assigned_cluster and word.custom_clusters.exists():
                assigned_cluster = word.custom_clusters.first().name
                
            result.append({
                'id': word.id,
                'word': word.word,
                'original_text': word.original_text,
                'frequency': word.frequency,
                'sentiment_score': word.sentiment_score,
                'assigned_cluster': assigned_cluster,
                'answer_id': word.answer_id
            })
            
        return DRFResponse(result)

    @action(detail=False, methods=['post'])
    @transaction.atomic
    def update_word_cluster(self, request):
        """
        Update the assigned cluster for a specific ResponseWord.
        """
        word_id = request.data.get('word_id')
        cluster_name = request.data.get('cluster_name')
        
        if not word_id or not cluster_name:
            return DRFResponse(
                {'detail': 'Both word_id and cluster_name are required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
            
        try:
            from .models import ResponseWord, CustomWordCluster
            word = ResponseWord.objects.get(id=word_id)
            
            # Update the directly assigned cluster
            word.assigned_cluster = cluster_name
            word.save()
            
            # Check if the cluster exists, if not create it
            cluster, created = CustomWordCluster.objects.get_or_create(
                name=cluster_name,
                defaults={
                    'created_by': request.user if request.user.is_authenticated else None,
                    'is_active': True
                }
            )
            
            # Associate the word with the cluster if it's not already
            if cluster not in word.custom_clusters.all():
                word.custom_clusters.add(cluster)
                
            return DRFResponse({
                'id': word.id,
                'word': word.word,
                'assigned_cluster': word.assigned_cluster,
                'detail': 'Cluster updated successfully'
            })
            
        except ResponseWord.DoesNotExist:
            return DRFResponse(
                {'detail': f'Word with id {word_id} not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return DRFResponse(
                {'detail': f'Error updating cluster: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    @transaction.atomic
    def submit_response(self, request):
        survey_id = request.data.get('survey')
        language = request.data.get('language', 'en')
        answers_data = request.data.get('answers', [])
        token = request.data.get('token')  # Get the token from the request
        
        # Log request data for debugging
        logger.info(f"Submit response request data: {request.data}")
        logger.info(f"Token received in request: {token}")
        
        try:
            survey = Survey.objects.get(id=survey_id)
        except Survey.DoesNotExist:
            return DRFResponse({'detail': 'Survey not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Check if the selected language is available for this survey
        if language not in survey.languages:
            return DRFResponse(
                {'detail': f'This survey is not available in {language}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if survey is active
        if not survey.is_active:
            return DRFResponse({'detail': 'This survey is no longer active'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if survey has expired
        if survey.expiry_date and survey.expiry_date < timezone.now():
            return DRFResponse({'detail': 'This survey has expired'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Generate a session ID if not provided
        session_id = request.data.get('session_id')
        if not session_id:
            import uuid
            session_id = str(uuid.uuid4())
        
        # Find SurveyToken if it exists
        survey_token = None
        if token:
            try:
                survey_token = SurveyToken.objects.get(token=token, survey=survey)
            except SurveyToken.DoesNotExist:
                # If not found in SurveyToken model, check if it matches the legacy token
                if survey.token == token:
                    # It's a legacy token
                    pass
                else:
                    # Invalid token
                    return DRFResponse({'detail': 'Invalid token'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Create response with token information
        response = Response.objects.create(
            survey=survey, 
            session_id=session_id, 
            language=language,
            token=token,
            survey_token=survey_token
        )
        
        # Process the answers
        required_questions = set(Question.objects.filter(survey=survey, is_required=True).values_list('id', flat=True))
        answered_required = set()
        
        logger.info(f"Processing {len(answers_data)} answers for response {response.id}")
        
        created_answers = []
        for answer_data in answers_data:
            question_id = answer_data.get('question')
            try:
                question = Question.objects.get(id=question_id, survey=survey)
                if question.is_required:
                    answered_required.add(question_id)
                
                answer = Answer.objects.create(
                    response=response,
                    question=question,
                    nps_rating=answer_data.get('nps_rating'),
                    text_answer=answer_data.get('text_answer')
                )
                created_answers.append(answer.id)
                logger.info(f"Created answer {answer.id} for question {question_id}")
            except Question.DoesNotExist:
                logger.warning(f"Question {question_id} not found for survey {survey.id}")
                pass
            except Exception as e:
                logger.error(f"Error creating answer for question {question_id}: {str(e)}")
                raise
        
        logger.info(f"Successfully created {len(created_answers)} answers for response {response.id}")
        
        # Check if all required questions were answered
        missing_questions = Question.objects.filter(id__in=required_questions - answered_required)
        if missing_questions.exists():
            response.delete()  # Clean up the incomplete response
            missing_texts = [f"{q.questions.get(language, q.questions.get('en', 'Untitled Question'))}" 
                           for q in missing_questions]
            return DRFResponse(
                {'detail': f'Please answer all required questions: {", ".join(missing_texts)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Stop here and return success - don't process answers synchronously
        return DRFResponse({
            'detail': 'Response submitted successfully',
            'response_id': response.id,
            'session_id': session_id
        })


class DashboardViewSet(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        # Get counts based on user role
        is_admin = request.user.groups.filter(name='Admin').exists()
        is_organizer = request.user.groups.filter(name='Organizer').exists()
        
        # Get survey counts
        if is_admin or is_organizer:
            # Admin and Organizer see all surveys
            total_surveys = Survey.objects.count()
            surveys = Survey.objects.all()
        else:
            # Others only see their own surveys
            total_surveys = Survey.objects.filter(created_by=request.user).count()
            surveys = Survey.objects.filter(created_by=request.user)
        
        # Get response counts for accessible surveys
        total_responses = Response.objects.filter(survey__in=surveys).count()
        
        # Calculate survey completion rate
        completion_rate = 0
        if total_surveys > 0:
            completed_surveys = 0
            for survey in surveys:
                if self.calculate_survey_completion(survey) >= 100:
                    completed_surveys += 1
            
            completion_rate = (completed_surveys / total_surveys) * 100
        
        # Recent activity - show responses for accessible surveys
        recent_responses = Response.objects.filter(
            survey__in=surveys
        ).order_by('-created_at')[:5]
        
        recent_activity = []
        for response in recent_responses:
            recent_activity.append({
                'type': 'response',
                'description': f"New response to {response.survey.title}",
                'date': response.created_at.strftime("%Y-%m-%d %H:%M")
            })
        
        # Prepare response data
        response_data = {
            'total_surveys': total_surveys,
            'total_responses': total_responses,
            'survey_completion_rate': round(completion_rate, 1),
            'recent_activity': recent_activity
        }
        
        # Only include user stats for Admin
        if is_admin:
            # Calculate user growth rate
            month_ago = timezone.now() - timedelta(days=30)
            total_users = User.objects.count()
            users_month_ago = User.objects.filter(date_joined__lt=month_ago).count()
            
            user_growth_rate = 0
            if users_month_ago > 0:
                user_growth_rate = ((total_users - users_month_ago) / users_month_ago) * 100
            
            response_data.update({
                'total_users': total_users,
                'user_growth_rate': round(user_growth_rate, 1)
            })
        
        return DRFResponse(response_data)
    
    def calculate_survey_completion(self, survey):
        """Calculate if a survey is complete based on required fields."""
        # Check if all required fields are filled in
        required_fields = [
            survey.title,
            survey.token
        ]
        
        # Count how many required fields are filled
        filled_fields = sum(1 for field in required_fields if field)
        
        # Check if the survey has at least one question
        has_questions = Question.objects.filter(survey=survey).exists()
        
        if not filled_fields:
            return 0
            
        # Calculate completion percentage
        completion = (filled_fields / len(required_fields)) * 100
        
        # If no questions, cap at 90%
        if not has_questions:
            completion = min(completion, 90)
            
        return completion


class SurveyAnalysisViewSet(viewsets.ViewSet):
    """
    ViewSet for advanced survey analysis features.
    """
    permission_classes = [permissions.IsAuthenticated, IsCreatorOrReadOnly]
    
    @action(detail=True, methods=['get'])
    def summary(self, request, pk=None):
        """Get analysis summary for a survey."""
        try:
            survey = Survey.objects.get(pk=pk)
            self.check_object_permissions(request, survey)
            
            # Get analysis summary, create if it doesn't exist
            summary, created = SurveyAnalysisSummary.objects.get_or_create(survey=survey)
            
            # Only update if it's a brand new summary
            if created:
                # Create a basic summary with response count and language breakdown
                from django.db.models import Count
                
                # Count responses
                response_count = Response.objects.filter(survey=survey).count()
                summary.response_count = response_count
                
                # Language breakdown
                languages = Response.objects.filter(survey=survey).values('language').annotate(
                    count=Count('id')
                )
                language_breakdown = {item['language']: item['count'] for item in languages}
                summary.language_breakdown = language_breakdown
                
                # Save the basic summary
                summary.save()
                
                # Log that a new summary was created
                logger.info(f"Created new analysis summary for survey {pk}")
            
            serializer = SurveyAnalysisSummarySerializer(summary)
            return DRFResponse(serializer.data)
            
        except Survey.DoesNotExist:
            return DRFResponse({'detail': 'Survey not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['get'])
    def word_cloud(self, request, pk=None):
        """Generate word cloud data for survey text responses including sentence context."""
        try:
            survey = Survey.objects.get(pk=pk)
            self.check_object_permissions(request, survey)
            
            # Get language parameter (default to first available language)
            language = request.query_params.get('language')
            if not language and survey.languages:
                language = survey.languages[0]
            
            # Get responses filtered by language
            responses = Response.objects.filter(survey=survey)
            if language:
                responses = responses.filter(language=language)
            
            # Get all text answers from these responses
            answers = Answer.objects.filter(
                response__in=responses,
                text_answer__isnull=False
            ).exclude(text_answer='')
            
            # Use ResponseWord data that already has sentence information
            from django.db.models import Count, Avg
            from .models import ResponseWord
            
            # Get the most frequent words for this survey and language
            word_data = ResponseWord.objects.filter(
                response__in=responses,
                language=language
            ).values('word').annotate(
                value=Count('id'),
                avg_sentiment=Avg('sentiment_score')
            ).order_by('-value')[:100]
            
            # Format data for word cloud with sentence information
            word_cloud_data = []
            
            for word_info in word_data:
                word = word_info['word']
                
                # Get all instances of this word to collect sentences
                word_instances = ResponseWord.objects.filter(
                    response__in=responses,
                    language=language,
                    word=word
                )
                
                # Get sentences from answer.sentence_sentiments for this word
                sentence_texts = []
                sentence_indices = []
                sentence_sentiments = []
                
                for word_instance in word_instances:
                    # Get the answer for this word instance
                    answer = word_instance.answer
                    
                    # Skip if no sentence_sentiments available
                    if not answer.sentence_sentiments:
                        continue
                    
                    # Get the sentence index from the word
                    sentence_idx = word_instance.sentence_index
                    
                    # If we have a valid sentence index, find the matching sentence
                    if sentence_idx is not None:
                        # Find the corresponding sentence in sentence_sentiments
                        for sent in answer.sentence_sentiments:
                            if sent.get('index') == sentence_idx:
                                sent_text = sent.get('text', '')
                                if sent_text and sent_text not in sentence_texts:
                                    sentence_texts.append(sent_text)
                                    sentence_indices.append(sentence_idx)
                                    sentence_sentiments.append(sent.get('sentiment', 0))
                
                # Find associated NPS score if available
                nps_scores = []
                print(f"Word: {word}, Instances: {word_instances.count()}")
                print(word_instances)
                for word_instance in word_instances:
                    nps_answer = Answer.objects.filter(
                        response=word_instance.response,
                        nps_rating__isnull=False
                    ).first()
                    if nps_answer and nps_answer.nps_rating is not None:
                        nps_scores.append(nps_answer.nps_rating)
                
                avg_nps = sum(nps_scores) / len(nps_scores) if nps_scores else None
                
                # Create the word cloud item with sentence context
                word_item = {
                    'text': word,
                    'value': word_info['value'],
                    'sentiment': sum(sentence_sentiments) / len(sentence_sentiments),
                    'sentence_texts': sentence_texts,
                    'sentence_indices': sentence_indices,
                    'sentence_sentiments': sentence_sentiments,
                    'nps_score': avg_nps
                }
                word_cloud_data.append(word_item)
            
            return DRFResponse(word_cloud_data)
            
        except Survey.DoesNotExist:
            return DRFResponse({'detail': 'Survey not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error generating word cloud: {str(e)}", exc_info=True)
            return DRFResponse(
                {"error": f"Error generating word cloud: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def cluster_cloud(self, request, pk=None):
        """Generate word cloud data for clusters in a survey, including sentence examples."""
        try:
            from django.db.models import Count, Avg
            from .models import Survey, ResponseWord, CustomWordCluster, Answer
            import logging
            
            logger = logging.getLogger(__name__)
            
            survey = Survey.objects.get(pk=pk)
            self.check_object_permissions(request, survey)
            
            # Get all custom clusters used in this survey's response words
            custom_clusters = CustomWordCluster.objects.filter(
                words__response__survey=survey
            ).distinct()
            
            # Log the number of clusters found
            logger.info(f"Found {custom_clusters.count()} custom clusters for survey {pk}")
            
            # Create a list to hold word cloud data
            cluster_cloud_data = []
            
            # For each custom cluster, collect statistics
            for cc in custom_clusters:
                # Get response words for this cluster
                response_words = ResponseWord.objects.filter(
                    custom_clusters=cc,
                    response__survey=survey
                )
                
                # Skip clusters with no words
                if not response_words.exists():
                    continue
                
                # Count distinct responses
                distinct_responses = response_words.values('response').distinct().count()
                
                # Include all clusters, even if they appear in only one response
                if distinct_responses < 1:
                    continue
                    
                # Calculate average sentiment
                avg_sentiment = response_words.aggregate(Avg('sentiment_score'))['sentiment_score__avg'] or 0
                
                # Get NPS ratings from associated responses
                nps_answers = Answer.objects.filter(
                    response__in=response_words.values('response').distinct(),
                    nps_rating__isnull=False
                )
                
                # Calculate average NPS and determine category
                avg_nps = None
                is_positive = False
                is_negative = False
                is_neutral = True
                
                
                
                # Collect all unique sentences from all words in this cluster
                all_sentences = []
                all_sentence_sentiments = []
                
                for word in response_words:
                    # Get the answer
                    answer = word.answer
                    
                    # Skip if no sentence_sentiments available
                    if not answer.sentence_sentiments:
                        continue
                    
                    # Get the sentence index from the word
                    sentence_idx = word.sentence_index
                    
                    # If we have a valid sentence index, find the matching sentence
                    if sentence_idx is not None:
                        # Find the corresponding sentence in sentence_sentiments
                        for sent in answer.sentence_sentiments:
                            if sent.get('index') == sentence_idx:
                                sent_text = sent.get('text', '')
                                if sent_text and sent_text not in all_sentences:
                                    all_sentences.append(sent_text)
                                    all_sentence_sentiments.append(sent.get('sentiment', 0))
                
                # Log cluster details to debug
                logger.info(f"Cluster: {cc.name}, Responses: {distinct_responses}, Sentiment: {avg_sentiment}, NPS: {avg_nps}")
                
                if nps_answers.exists():
                    avg_sentiment = sum(all_sentence_sentiments) / len(all_sentence_sentiments)
                    nps_scores = list(nps_answers.values_list('nps_rating', flat=True))
                    avg_nps = sum(nps_scores) / len(nps_scores) if nps_scores else None
                    # print("Avg Sentiment: " + str(avg_sentiment))
                    if avg_sentiment is not None:
                        if avg_sentiment > 0:
                            is_positive = True
                            is_neutral = False
                        elif avg_sentiment < 0:
                            is_negative = True
                            is_neutral = False

                elif avg_sentiment > 0.2:
                    is_positive = True
                    is_neutral = False
                elif avg_sentiment < -0.2:
                    is_negative = True
                    is_neutral = False
                # Add to word cloud data
                cluster_cloud_data.append({
                    'text': cc.name,
                    'value': distinct_responses,
                    'sentiment': avg_sentiment,
                    'is_positive': is_positive,
                    'is_negative': is_negative,
                    'is_neutral': is_neutral,
                    'nps_score': avg_nps,
                    'sentences': all_sentences,
                    'sentence_sentiments': all_sentence_sentiments,
                    'keywords': cc.keywords[:5] if cc.keywords else [],
                    'total_words': response_words.count()
                })
            
            # Sort by frequency (value)
            cluster_cloud_data.sort(key=lambda x: x['value'], reverse=True)
            
            # Log the number of clusters in the cloud data
            logger.info(f"Returning {len(cluster_cloud_data)} clusters for word cloud")
            
            # Return all clusters, don't limit to a specific number
            return DRFResponse(cluster_cloud_data)
            
        except Survey.DoesNotExist:
            return DRFResponse({'detail': 'Survey not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error generating cluster cloud: {str(e)}", exc_info=True)
            return DRFResponse(
                {"error": f"Error generating cluster cloud: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def clusters(self, request, pk=None):
        """
        Get clusters for a survey, optionally filtered by type (positive, negative, neutral).
        Uses CustomWordClusters from ResponseWord instances.
        """
        try:
            from django.db.models import Count, Avg, Q
            from .models import Survey, ResponseWord, CustomWordCluster, Answer
            
            survey = Survey.objects.get(pk=pk)
            self.check_object_permissions(request, survey)
            
            # Get all custom clusters used in this survey's response words
            custom_clusters = CustomWordCluster.objects.filter(
                words__response__survey=survey
            ).distinct()
            
            # Create a list to hold cluster data
            cluster_data_list = []
            
            # For each custom cluster, collect statistics
            for cc in custom_clusters:
                # Get response words for this cluster
                response_words = ResponseWord.objects.filter(
                    custom_clusters=cc,
                    response__survey=survey
                )
                
                # Skip clusters with no words
                if not response_words.exists():
                    continue
                
                # Count unique sentences to avoid multiple counting
                unique_sentences = set()
                for word in response_words:
                    if word.sentence_index is not None:
                        unique_sentences.add((word.response_id, word.answer_id, word.sentence_index))
                
                # Frequency is the number of unique sentences
                frequency = len(unique_sentences)
                
                # Calculate average sentiment
                avg_sentiment = response_words.aggregate(Avg('sentiment_score'))['sentiment_score__avg'] or 0
                
                # Get NPS ratings from associated responses
                nps_answers = Answer.objects.filter(
                    response__in=response_words.values('response').distinct(),
                    nps_rating__isnull=False
                )
                
                # Calculate average NPS and determine category
                avg_nps = None
                is_positive = False
                is_negative = False
                is_neutral = True
                
                if nps_answers.exists():
                    nps_scores = list(nps_answers.values_list('nps_rating', flat=True))
                    avg_nps = sum(nps_scores) / len(nps_scores) if nps_scores else None
                    
                    if avg_nps is not None:
                        if avg_nps >= 9:
                            is_positive = True
                            is_neutral = False
                        elif avg_nps <= 6:
                            is_negative = True
                            is_neutral = False
                elif avg_sentiment > 0.2:
                    is_positive = True
                    is_neutral = False
                elif avg_sentiment < -0.2:
                    is_negative = True
                    is_neutral = False
                
                # Create cluster data object
                cluster_data = {
                    'id': cc.id,
                    'name': cc.name,
                    'survey': survey.id,
                    'description': cc.description or '',
                    'frequency': frequency,
                    'sentiment_score': avg_sentiment,
                    'nps_score': avg_nps,
                    'is_positive': is_positive,
                    'is_negative': is_negative,
                    'is_neutral': is_neutral,
                    'custom_cluster_id': cc.id,
                    'created_at': cc.created_at.isoformat() if cc.created_at else None,
                    'updated_at': cc.updated_at.isoformat() if cc.updated_at else None
                }
                
                cluster_data_list.append(cluster_data)
            
            # Filter by type if requested - ensure strict filtering by boolean values
            cluster_type = request.query_params.get('type')
            if cluster_type == 'positive':
                cluster_data_list = [c for c in cluster_data_list if c['is_positive'] == True]
            elif cluster_type == 'negative':
                cluster_data_list = [c for c in cluster_data_list if c['is_negative'] == True]
            elif cluster_type == 'neutral':
                cluster_data_list = [c for c in cluster_data_list if c['is_neutral'] == True]
            
            # Sort by frequency
            cluster_data_list.sort(key=lambda x: x['frequency'], reverse=True)
            
            # Limit results
            limit = int(request.query_params.get('limit', 10))
            cluster_data_list = cluster_data_list[:limit]
            
            return DRFResponse(cluster_data_list)
            
        except Survey.DoesNotExist:
            return DRFResponse({'detail': 'Survey not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error getting clusters: {str(e)}", exc_info=True)
            return DRFResponse(
                {"error": f"Error getting clusters: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _analyze_survey_responses(self, survey):
        """Analyze all responses for a survey and extract insights."""
        # Get all responses for the survey
        responses = Response.objects.filter(survey=survey)
        
        # print("Response length")
        # print(len(responses))
        # Process each response
        for response in responses:
            self._analyze_single_response(response)
        
        # After processing all responses, identify clusters
        self._generate_word_clusters(survey)
    
    def _analyze_single_response(self, response):
        """Analyze a single response to extract words and sentiments."""
        # Skip responses that have already been processed
        if ResponseWord.objects.filter(response=response).exists():
            return
        
        # Get all text answers from this response
        text_answers = Answer.objects.filter(
            response=response,
            text_answer__isnull=False
        ).exclude(text_answer='')
        
        analyzer = TextAnalyzer(language=response.language)
        
        # Process each text answer
        for answer in text_answers:
            if not answer.text_answer:
                continue
                
            # Calculate sentiment for the answer if not already done
            if answer.sentiment_score is None:
                answer.sentiment_score = analyzer.get_sentiment_score(answer.text_answer)
                answer.save()
            
            # Extract words from the answer
            word_freq = analyzer.get_word_frequencies(answer.text_answer)
            
            # Save each word with its sentiment and frequency
            for word, frequency in word_freq.items():
                # Calculate sentiment specifically for this word in context
                word_sentiment = analyzer.get_word_sentiment(word, answer.text_answer)
                
                # Create or update the ResponseWord
                ResponseWord.objects.create(
                    response=response,
                    answer=answer,
                    word=word,
                    original_text=answer.text_answer,
                    frequency=frequency,
                    sentiment_score=word_sentiment,
                    language=response.language
                )
    
    def _generate_word_clusters(self, survey):
        """
        Calculate cluster metrics from existing CustomWordCluster assignments for SurveyAnalysisSummary.
        This does NOT create new WordCluster objects, but uses the existing clusters already 
        assigned to ResponseWord model instances.
        """
        logger.info(f"Analyzing existing custom clusters for survey {survey.id}")
        
        # Get all custom clusters used in this survey's response words
        custom_clusters = CustomWordCluster.objects.filter(
            words__response__survey=survey
        ).distinct()
        
        if not custom_clusters.exists():
            logger.warning(f"No custom clusters found for survey {survey.id}")
            return
            
        logger.info(f"Found {custom_clusters.count()} custom clusters for survey {survey.id}")
        
        # Create a list to hold cluster data
        cluster_data = []
        
        # Process each custom cluster
        for cc in custom_clusters:
            # Get response words for this cluster
            response_words = ResponseWord.objects.filter(
                custom_clusters=cc,
                response__survey=survey
            )
            
            # Skip clusters with no words
            if not response_words.exists():
                continue

            # print(cc)
            # print(response_words)
            
            # Count unique responses where this cluster appears
            # This is a more reliable measure than sentence count
            distinct_responses = response_words.values('response').distinct().count()
            
            # As a fallback, also count unique sentences
            unique_sentences = set()
            sentence_sentiment_scores = []  # Collect sentiment scores from sentences
            
            # Track which sentences we've already processed to avoid duplicates
            processed_sentences = set()
            
            for word in response_words:
                # We need both answer and sentence_index to find the sentence sentiment
                if word.answer_id and word.sentence_index is not None:
                    sentence_key = (word.answer_id, word.sentence_index)
                    
                    # Keep track of unique sentences
                    unique_sentences.add((word.response_id, word.answer_id, word.sentence_index))
                    
                    # Only process each sentence once for sentiment calculation
                    if sentence_key not in processed_sentences:
                        processed_sentences.add(sentence_key)
                        
                        # Get the sentiment from the sentence_sentiments array in the answer
                        if word.answer.sentence_sentiments:
                            for sent in word.answer.sentence_sentiments:
                                if sent.get('index') == word.sentence_index:
                                    # Found the matching sentence, add its sentiment to our list
                                    sentiment_score = sent.get('sentiment', 0)
                                    sentence_sentiment_scores.append(sentiment_score)
                                    break
            
            # Use response count as primary frequency, fallback to unique sentences if no responses
            frequency = max(distinct_responses, len(unique_sentences))
            
            # Calculate average sentiment based on sentence sentiments
            # This is more accurate than using word sentiment scores
            avg_sentiment = 0
            if sentence_sentiment_scores:
                avg_sentiment = sum(sentence_sentiment_scores) / len(sentence_sentiment_scores)
            
            # Get NPS ratings from associated responses - get ALL NPS answers from these responses
            response_ids = response_words.values_list('response_id', flat=True).distinct()
            nps_answers = Answer.objects.filter(
                response_id__in=response_ids,
                question__type='nps',
                nps_rating__isnull=False
            )
            
            # Calculate average NPS with error handling
            avg_nps = None
            if nps_answers.exists():
                try:
                    nps_scores = list(nps_answers.values_list('nps_rating', flat=True))
                    if nps_scores:
                        # Filter out None values and handle empty lists
                        nps_scores = [score for score in nps_scores if score is not None]
                        avg_nps = sum(nps_scores) / len(nps_scores) if nps_scores else None
                except Exception as e:
                    logger.error(f"Error calculating NPS for cluster {cc.name}: {str(e)}")
            
            # Determine cluster sentiment category with improved thresholds
            is_positive = False
            is_negative = False
            is_neutral = True
                
            if avg_nps is not None:
                if avg_nps >= 9:
                    is_positive = True
                    is_neutral = False
                elif avg_nps <= 6:
                    is_negative = True
                    is_neutral = False
            elif avg_sentiment > 0.25:  # Slightly higher threshold for positive sentiment
                is_positive = True
                is_neutral = False
            elif avg_sentiment < -0.25:  # Slightly lower threshold for negative sentiment
                is_negative = True
                is_neutral = False
            
            # Collect data for this cluster with the improved metrics
            cluster_data.append({
                'id': cc.id,
                'name': cc.name,
                'description': cc.description,
                'frequency': frequency,
                'response_count': distinct_responses,  # Store actual response count separately
                'sentence_count': len(unique_sentences),  # Store sentence count separately
                'sentiment_score': avg_sentiment,
                'is_positive': is_positive,
                'is_negative': is_negative,
                'is_neutral': is_neutral,
                'nps_score': avg_nps,
            })
        # print(cluster_data)
        # Get or create the survey analysis summary
        summary, _ = SurveyAnalysisSummary.objects.get_or_create(survey=survey)
        
        # Sort by frequency and update top clusters - now using the improved frequency metric
        all_clusters = sorted(cluster_data, key=lambda x: x['frequency'], reverse=True)
        positive_clusters = sorted([c for c in cluster_data if c['is_positive']], 
                                  key=lambda x: x['frequency'], reverse=True)
        negative_clusters = sorted([c for c in cluster_data if c['is_negative']], 
                                  key=lambda x: x['frequency'], reverse=True)
        neutral_clusters = sorted([c for c in cluster_data if c['is_neutral']], 
                                 key=lambda x: x['frequency'], reverse=True)
        
        # Update the summary with cluster IDs
        summary.top_clusters = [c['id'] for c in all_clusters]
        summary.top_positive_clusters = [c['id'] for c in positive_clusters]
        summary.top_negative_clusters = [c['id'] for c in negative_clusters]
        summary.top_neutral_clusters = [c['id'] for c in neutral_clusters]
        # print(all_clusters)
        # print(summary.top_clusters)
        # Store additional metrics in the summary
        summary.metrics = {
            'cluster_metrics': {str(c['id']): {  # Convert ID to string since JSON keys must be strings
                'name': c['name'],
                'description': c.get('description', ''),
                'frequency': c['frequency'],
                'response_count': c['response_count'],
                'sentence_count': c['sentence_count'],
                'sentiment_score': c['sentiment_score'],
                'is_positive': c['is_positive'],
                'is_negative': c['is_negative'], 
                'is_neutral': c['is_neutral'],
                'nps_score': c['nps_score']
            } for c in cluster_data}
        }
        
        # CALCULATE SATISFACTION METRICS
        # Get all NPS ratings from survey responses
        import statistics
        import math
        from .utils import calculate_satisfaction_score
        
        # Get all NPS ratings for the survey
        nps_ratings = list(Answer.objects.filter(
            response__survey=survey,
            question__type='nps',
            nps_rating__isnull=False
        ).values_list('nps_rating', flat=True))
        
        # Filter out None values
        nps_ratings = [score for score in nps_ratings if score is not None]
        
        if nps_ratings:
            # Calculate average satisfaction (average NPS score)
            summary.average_satisfaction = sum(nps_ratings) / len(nps_ratings)
            
            # Calculate median satisfaction
            summary.median_satisfaction = statistics.median(nps_ratings)
            
            # Calculate satisfaction confidence interval (95%)
            # Using the formula: mean ± 1.96 * (standard_deviation / sqrt(n))
            if len(nps_ratings) > 1:  # Need at least 2 values for std dev
                std_dev = statistics.stdev(nps_ratings)
                margin_of_error = 1.96 * (std_dev / math.sqrt(len(nps_ratings)))
                summary.satisfaction_confidence_low = max(0, summary.average_satisfaction - margin_of_error)
                summary.satisfaction_confidence_high = min(10, summary.average_satisfaction + margin_of_error)
            else:
                # Just one rating, set confidence to that value
                summary.satisfaction_confidence_low = summary.average_satisfaction
                summary.satisfaction_confidence_high = summary.average_satisfaction
            
            # Calculate NPS score using NPS methodology (% promoters - % detractors)
            satisfaction_data = calculate_satisfaction_score(nps_ratings)
            summary.satisfaction_score = satisfaction_data['score']
            
            # Calculate sentiment divergence
            if len(nps_ratings) > 1:
                summary.sentiment_divergence = std_dev  # Use standard deviation as measure of divergence
            else:
                summary.sentiment_divergence = 0
            
            # Update percentages based on NPS categories
            summary.positive_percentage = satisfaction_data['promoters_pct']
            summary.negative_percentage = satisfaction_data['detractors_pct']
            summary.neutral_percentage = satisfaction_data['passives_pct']
        else:
            # If no NPS ratings, try to calculate from sentence sentiments
            all_sentiments = []
            for answer in Answer.objects.filter(
                response__survey=survey,
                text_answer__isnull=False,
                processed=True
            ):
                if answer.sentence_sentiments:
                    for sent in answer.sentence_sentiments:
                        if 'sentiment' in sent:
                            all_sentiments.append(sent['sentiment'])
            
            if all_sentiments:
                # Calculate averages and percentages based on sentiments
                summary.average_satisfaction = sum(all_sentiments) / len(all_sentiments) * 5 + 5  # Scale -1..1 to 0..10
                summary.median_satisfaction = (statistics.median(all_sentiments) * 5) + 5
                
                positive_count = sum(1 for s in all_sentiments if s > 0.05)
                negative_count = sum(1 for s in all_sentiments if s < -0.05)
                neutral_count = len(all_sentiments) - positive_count - negative_count
                
                total = len(all_sentiments)
                summary.positive_percentage = (positive_count / total) * 100 if total > 0 else 0
                summary.negative_percentage = (negative_count / total) * 100 if total > 0 else 0
                summary.neutral_percentage = (neutral_count / total) * 100 if total > 0 else 0
                
                # Calculate satisfaction score as (positive % - negative %)
                summary.satisfaction_score = summary.positive_percentage - summary.negative_percentage
                
                # Calculate confidence interval
                if len(all_sentiments) > 1:
                    std_dev = statistics.stdev(all_sentiments)
                    margin_of_error = 1.96 * (std_dev / math.sqrt(len(all_sentiments)))
                    scaled_margin = margin_of_error * 5  # Scale to our 0-10 scale
                    summary.satisfaction_confidence_low = max(0, summary.average_satisfaction - scaled_margin)
                    summary.satisfaction_confidence_high = min(10, summary.average_satisfaction + scaled_margin)
                    summary.sentiment_divergence = std_dev
                else:
                    summary.satisfaction_confidence_low = summary.average_satisfaction
                    summary.satisfaction_confidence_high = summary.average_satisfaction
                    summary.sentiment_divergence = 0
        
        # Update response count
        summary.response_count = Response.objects.filter(survey=survey).count()
        
        # Language breakdown
        from django.db.models import Count
        languages = Response.objects.filter(survey=survey).values('language').annotate(
            count=Count('id')
        )
        language_breakdown = {item['language']: item['count'] for item in languages}
        summary.language_breakdown = language_breakdown
        
        # Save the updated summary
        summary.save()
        
        logger.info(f"Updated analysis summary with {len(all_clusters)} clusters and satisfaction metrics for survey {survey.id}")
        
        # No need to update any WordCluster objects - we're using the existing CustomWordClusters
    
    def _update_analysis_summary(self, summary):
        """
        Update the analysis summary with top clusters.
        This now redirects to _generate_word_clusters for complete analysis.
        """
        logger.info(f"Redirecting to _generate_word_clusters for complete analysis summary update")
        if hasattr(summary, 'survey'):
            self._generate_word_clusters(summary.survey)
        return summary

    @action(detail=True, methods=['post'])
    def process_all_responses(self, request, pk=None):
        """
        Process all responses for a survey via direct method.
        This is a simpler implementation that directly processes responses
        without the complexity of the analyze_responses method.
        """
        try:
            survey = Survey.objects.get(pk=pk)
            self.check_object_permissions(request, survey)
            
            from .direct_process_all_responses import direct_process_all_responses
            result = direct_process_all_responses(survey.id)
            
            if result['success']:
                return DRFResponse({
                                'detail': f"Successfully processed {result['processed_count']} responses for survey {survey.id}",
                                'processed_count': result['processed_count'],
                                'total_responses': result['total_responses'],
                                'message': f"Successfully processed {result['processed_count']} out of {result['total_responses']} responses.",
                                'cluster_count': result.get('cluster_count', 0),
                                'clusters': result.get('clusters', [])
                            })
            else:
                return DRFResponse({
                    'detail': f"Error processing survey: {result['error']}",
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
        except Survey.DoesNotExist:
            return DRFResponse({'detail': 'Survey not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error in process_all_responses: {str(e)}", exc_info=True)
            return DRFResponse({'detail': f"Error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def analyze_responses(self, request, pk=None):
        """
        Analyze survey responses to update the analysis summary using existing custom clusters.
        This does NOT generate any new clusters, but uses the existing custom cluster assignments
        from ResponseWord model instances.
        """
        try:
            # Get the survey by ID directly instead of using get_object()
            survey = Survey.objects.get(pk=pk)
            
            # Check permissions
            self.check_object_permissions(request, survey)
            
            logger.info(f"Starting analysis for survey {survey.id}")
            
            # Check if custom clusters exist for this survey
            has_custom_clusters = ResponseWord.objects.filter(
                response__survey=survey,
                custom_clusters__isnull=False
            ).exists()
            
            if not has_custom_clusters:
                logger.warning(f"No custom clusters found for survey {survey.id}")
                return DRFResponse({
                    "status": "error",
                    "message": "No custom clusters found for this survey's responses."
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Generate cluster metrics from existing custom cluster assignments
            self._generate_word_clusters(survey)
            
            # Get the latest summary data 
            summary, _ = SurveyAnalysisSummary.objects.get_or_create(survey=survey)
            
            # Count the responses
            response_count = Response.objects.filter(survey=survey).count()
            
            # Count the custom clusters
            cluster_count = CustomWordCluster.objects.filter(
                words__response__survey=survey
            ).distinct().count()
            
            logger.info(f"Analysis completed for survey {survey.id}, {response_count} responses and {cluster_count} clusters")
            
            # Prepare satisfaction metrics for response
            satisfaction_metrics = {
                "average_satisfaction": summary.average_satisfaction,
                "median_satisfaction": summary.median_satisfaction,
                "satisfaction_confidence_low": summary.satisfaction_confidence_low,
                "satisfaction_confidence_high": summary.satisfaction_confidence_high,
                "satisfaction_score": summary.satisfaction_score,
                "sentiment_divergence": summary.sentiment_divergence,
                "positive_percentage": summary.positive_percentage,
                "negative_percentage": summary.negative_percentage,
                "neutral_percentage": summary.neutral_percentage
            }
            
            return DRFResponse({
                "status": "success",
                "message": f"Analyzed {response_count} responses using {cluster_count} custom clusters",
                "data": {
                    "summary_id": summary.id,
                    "response_count": response_count,
                    "cluster_count": cluster_count,
                    "satisfaction_metrics": satisfaction_metrics
                }
            })
        except Survey.DoesNotExist:
            return DRFResponse({
                "status": "error",
                "message": "Survey not found"
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error analyzing responses for survey {pk}: {str(e)}")
            return DRFResponse({
                "status": "error",
                "message": f"Failed to analyze responses: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CustomWordClusterViewSet(viewsets.ModelViewSet):
    """
    API endpoint for managing custom word clusters.
    These are user-defined clusters that can be used across all surveys.
    """
    queryset = CustomWordCluster.objects.all()
    serializer_class = CustomWordClusterSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Filter clusters to only those created by the current user or shared/global clusters."""
        user = self.request.user
        
        # Admin can see all clusters
        if user.is_staff or user.groups.filter(name="Admin").exists():
            return CustomWordCluster.objects.all()
        
        # Others can only see their own clusters
        return CustomWordCluster.objects.filter(created_by=user)
    
    def perform_create(self, serializer):
        """Save the current user as the creator of the cluster."""
        serializer.save(created_by=self.request.user)
    
    def update(self, request, *args, **kwargs):
        """Handle PUT and PATCH requests to update a cluster, including keywords."""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        # First, let the serializer handle regular fields
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        
        # Update word count after updating keywords
        instance.word_count = len(instance.keywords or [])
        for lang, lang_keywords in (instance.multilingual_keywords or {}).items():
            instance.word_count += len(lang_keywords)
        instance.save(update_fields=['word_count'])
        
        # Return the updated instance
        return DRFResponse(serializer.data)
    
    @action(detail=False, methods=['get'])
    def active(self, request):
        """Get only active custom word clusters."""
        queryset = self.get_queryset().filter(is_active=True)
        serializer = self.get_serializer(queryset, many=True)
        return DRFResponse(serializer.data)
    
    @action(detail=True, methods=['post'])
    def toggle_status(self, request, pk=None):
        """Toggle the active status of a custom word cluster."""
        cluster = self.get_object()
        cluster.is_active = not cluster.is_active
        cluster.save()
        serializer = self.get_serializer(cluster)
        return DRFResponse(serializer.data)
    
    @action(detail=True, methods=['post'])
    def add_keywords(self, request, pk=None):
        """
        Add keywords to a custom word cluster.
        If language is specified, add to language-specific keywords.
        Otherwise, add to the legacy keywords list.
        """
        cluster = self.get_object()
        
        keywords = request.data.get('keywords', [])
        language = request.data.get('language', None)
        
        if not isinstance(keywords, list):
            keywords = [keywords]
            
        # Skip empty keywords
        keywords = [k for k in keywords if k and k.strip()]
        if not keywords:
            return DRFResponse({"detail": "No valid keywords provided"}, status=status.HTTP_400_BAD_REQUEST)
            
        # If language is specified, add to multilingual_keywords
        if language:
            # Initialize the language key if not present
            if not cluster.multilingual_keywords:
                cluster.multilingual_keywords = {}
                
            if language not in cluster.multilingual_keywords:
                cluster.multilingual_keywords[language] = []
                
            # Add new keywords, avoiding duplicates
            current_keywords = set(cluster.multilingual_keywords[language])
            current_keywords.update(keywords)
            cluster.multilingual_keywords[language] = list(current_keywords)
        else:
            # Add to legacy keywords list
            current_keywords = set(cluster.keywords or [])
            current_keywords.update(keywords)
            cluster.keywords = list(current_keywords)
        
        cluster.word_count = len(cluster.keywords or [])
        for lang, lang_keywords in (cluster.multilingual_keywords or {}).items():
            cluster.word_count += len(lang_keywords)
            
        cluster.save()
        
        # Return updated cluster
        serializer = self.get_serializer(cluster)
        return DRFResponse(serializer.data)
        
    @action(detail=True, methods=['post'])
    def remove_keyword(self, request, pk=None):
        """
        Remove a keyword from a custom word cluster.
        If language is specified, remove from language-specific keywords.
        Otherwise, remove from the legacy keywords list.
        """
        cluster = self.get_object()
        
        keyword = request.data.get('keyword')
        language = request.data.get('language', None)
        
        if not keyword:
            return DRFResponse({"detail": "No keyword provided"}, status=status.HTTP_400_BAD_REQUEST)
            
        if language:
            # Remove from multilingual_keywords if language is specified
            if cluster.multilingual_keywords and language in cluster.multilingual_keywords:
                if keyword in cluster.multilingual_keywords[language]:
                    cluster.multilingual_keywords[language].remove(keyword)
                    # Remove empty language lists
                    if not cluster.multilingual_keywords[language]:
                        del cluster.multilingual_keywords[language]
        else:
            # Remove from legacy keywords
            if cluster.keywords and keyword in cluster.keywords:
                cluster.keywords.remove(keyword)
                
        # Update word count
        cluster.word_count = len(cluster.keywords or [])
        for lang, lang_keywords in (cluster.multilingual_keywords or {}).items():
            cluster.word_count += len(lang_keywords)
            
        cluster.save()
        
        # Return updated cluster
        serializer = self.get_serializer(cluster)
        return DRFResponse(serializer.data)


class ProcessTextView(APIView):
    """
    API endpoint for testing text processing.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request, format=None):
        text = request.data.get('text', '')
        language = request.data.get('language', 'en')
        use_openai = request.data.get('use_openai', True)  # Default to using OpenAI
        
        if not text:
            return DRFResponse(
                {"error": "Text is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
            
        # Analyze sentences for sentiment using OpenAI or NLTK based on the request
        if use_openai:
            from .utils import analyze_sentences_with_openai
            sentence_data = analyze_sentences_with_openai(text, language)
        else:
            from .utils import analyze_sentences
            sentence_data = analyze_sentences(text, language)
        
        # Process words from each sentence
        from .utils import process_sentence
        all_processed_words = []
        words_to_sentences = {}
        
        # Process each sentence to extract words
        for sentence_info in sentence_data:
            sentence_text = sentence_info['text']
            sentence_idx = sentence_info['index']
            
            # Extract words from this sentence
            sentence_words = process_sentence(sentence_text, language)
            
            # Map each word to its source sentence
            for word in sentence_words:
                words_to_sentences[word] = {
                    'text': sentence_text,
                    'index': sentence_idx
                }
            
            # Add to our complete list of processed words
            all_processed_words.extend(sentence_words)
        
        # Remove duplicates while preserving order
        processed_words = []
        seen = set()
        for word in all_processed_words:
            if word not in seen:
                seen.add(word)
                processed_words.append(word)
        
        # Assign clusters to words without saving to database
        word_clusters = assign_clusters_to_words(text, processed_words, language)
        
        # Prepare structured result with word and assigned cluster
        structured_words = []
        for word in processed_words:
            structured_words.append({
                'word': word,
                'assigned_cluster': word_clusters.get(word, 'Other'),
                'sentence_index': words_to_sentences.get(word, {}).get('index')
            })
        
        # Return the processed words with cluster assignments and sentence data
        return DRFResponse({
            'original_text': text,
            'language': language,
            'processed_words': processed_words,
            'structured_words': structured_words,
            'sentences': sentence_data,
            'word_count': len(processed_words)
        }, status=status.HTTP_200_OK)


class ProcessSurveyResponsesView(APIView):
    """
    API endpoint for processing all responses for a survey and assigning clusters to words.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request, format=None):
        survey_id = request.data.get('survey_id')
        response_id = request.data.get('response_id')
        
        if not survey_id and not response_id:
            return DRFResponse(
                {"error": "Either survey_id or response_id is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
            
        try:
            if response_id:
                # Process a single response
                process_survey_and_assign_clusters(response_id)
                return DRFResponse({
                    'message': f'Successfully processed response {response_id}'
                }, status=status.HTTP_200_OK)
            else:
                # Process all responses for a survey
                from .models import Response, Survey
                
                try:
                    survey = Survey.objects.get(id=survey_id)
                except Survey.DoesNotExist:
                    return DRFResponse(
                        {"error": f"Survey with ID {survey_id} does not exist"}, 
                        status=status.HTTP_404_NOT_FOUND
                    )
                
                # Get all responses for this survey
                responses = Response.objects.filter(
                    answers__question__survey=survey
                ).distinct()
                
                # Process each response
                processed_count = 0
                for response in responses:
                    try:
                        process_survey_and_assign_clusters(response.id)
                        processed_count += 1
                    except Exception as e:
                        logger.error(f"Error processing response {response.id}: {str(e)}")
                
                return DRFResponse({
                    'message': f'Successfully processed {processed_count} responses for survey {survey_id}'
                }, status=status.HTTP_200_OK)

                
        except Exception as e:
            return DRFResponse(
                {"error": f"An error occurred: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class TemplateViewSet(viewsets.ModelViewSet):
    """
    API endpoint for managing templates.
    Templates are used to create surveys with predefined structures and clusters.
    """
    queryset = Template.objects.all()
    permission_classes = [permissions.IsAuthenticated, IsCreatorOrReadOnly]
    
    def get_serializer_class(self):
        if self.action in ['retrieve', 'update', 'partial_update', 'create']:
            return TemplateDetailSerializer
        return TemplateSerializer
    
    def get_queryset(self):
        user = self.request.user
        
        # Admin sees all templates
        if user.is_staff or user.is_superuser or user.groups.filter(name='Admin').exists():
            return Template.objects.all()
        
        # Others see only their templates
        return Template.objects.filter(created_by=user)
    
    @action(detail=True, methods=['post'])
    def add_cluster(self, request, pk=None):
        template = self.get_object()
        cluster_id = request.data.get('cluster_id')
        
        try:
            cluster = CustomWordCluster.objects.get(id=cluster_id)
            
            # Add the cluster to the template
            template.clusters.add(cluster)
            
            return Response({
                'status': 'success',
                'message': f'Cluster "{cluster.name}" added to template.'
            }, status=status.HTTP_200_OK)
            
        except CustomWordCluster.DoesNotExist:
            return Response({
                'status': 'error',
                'message': 'Cluster does not exist.'
            }, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['post'])
    def remove_cluster(self, request, pk=None):
        template = self.get_object()
        cluster_id = request.data.get('cluster_id')
        
        try:
            cluster = CustomWordCluster.objects.get(id=cluster_id)
            
            # Remove the cluster from the template
            template.clusters.remove(cluster)
            
            return Response({
                'status': 'success',
                'message': f'Cluster "{cluster.name}" removed from template.'
            }, status=status.HTTP_200_OK)
            
        except CustomWordCluster.DoesNotExist:
            return Response({
                'status': 'error',
                'message': 'Cluster does not exist.'
            }, status=status.HTTP_404_NOT_FOUND)

    def update(self, request, *args, **kwargs):
        """Override update method to add debugging"""
        print("\n\n=== TEMPLATE UPDATE DEBUG ===")
        print(f"Template ID: {kwargs.get('pk')}")
        print(f"Request data: {request.data}")
        if 'clusters' in request.data:
            print(f"Clusters in request: {request.data['clusters']}")
            print(f"Clusters type: {type(request.data['clusters'])}")
            
        response = super().update(request, *args, **kwargs)
        
        # Log the result for debugging
        template = self.get_object()
        print(f"After update - Template clusters: {list(template.clusters.values('id', 'name'))}")
        return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def survey_sentence_sentiment_analysis(request, survey_id):
    """
    Get detailed sentence-level sentiment analysis for a survey.
    
    This endpoint provides comprehensive sentiment analysis at the sentence level,
    including sentiment distribution, top positive and negative sentences,
    sentiment by questions, and sentiment by clusters.
    """
    # Get the survey or return 404
    survey = get_object_or_404(Survey, id=survey_id)
    
    # Check if the user has permission to access this survey
    if not request.user.is_staff and survey.created_by != request.user:
        return DRFResponse({"error": "You don't have permission to access this survey"}, status=403)
    
    # Get the analysis data
    analysis_data = get_survey_sentence_sentiment_analysis(survey)
    
    # Return the analysis data
    return DRFResponse(analysis_data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def process_answer_sentiment(request, answer_id):
    """
    Process or reprocess an existing answer to analyze sentence-level sentiment.
    
    This is useful for updating existing answers with the new sentence sentiment 
    analysis without reprocessing all answers.
    """
    from .models import Answer
    from .utils import analyze_sentences, process_sentence
    
    # Get the answer or return 404
    answer = get_object_or_404(Answer, id=answer_id)
    
    # Check if the user has permission to access this answer
    if not request.user.is_staff and answer.response.survey.created_by != request.user:
        return DRFResponse({"error": "You don't have permission to access this answer"}, status=403)
    
    # Skip if there's no text answer
    if not answer.text_answer:
        return DRFResponse({"error": "No text answer to process"}, status=400)
    
    # Get the language from the response
    language = answer.response.language
    
    # Analyze text at sentence level for sentiment
    sentence_data = analyze_sentences(answer.text_answer, language)
    answer.sentence_sentiments = sentence_data
    
    # If this is a new answer that hasn't been processed yet, fully process it
    if not answer.processed:
        # Process the answer fully
        answer.process_text_answer()
        return DRFResponse({"message": "Answer fully processed with sentence sentiment analysis"})
    else:
        # Just update the sentence sentiment data
        answer.save(update_fields=['sentence_sentiments'])
        
        # Additional step: Update sentence information for existing ResponseWord objects
        from .models import ResponseWord
        
        # Initialize mapping of words to sentences
        words_to_sentences = {}
        
        # Process each sentence to extract words and map them to sentences
        for idx, sentence_info in sentence_data:
            sentence_text = sentence_info['text']
            sentence_idx = sentence_info['index']
            
            # Extract words from this sentence
            sentence_words = process_sentence(sentence_text, language)
            
            # Map each word to its source sentence
            for word in sentence_words:
                words_to_sentences[word] = {
                    'text': sentence_text,
                    'index': sentence_idx
                }
        
        # Get all response words for this answer
        response_words = ResponseWord.objects.filter(answer=answer)
        
        # Update sentence information for each word
        for word in response_words:
            # Check if this word exists in our mapping
            if word.word in words_to_sentences:
                sentence_data = words_to_sentences[word.word]
                word.sentence_text = sentence_data['text']
                word.sentence_index = sentence_data['index']
                word.save(update_fields=['sentence_text', 'sentence_index'])
        
        return DRFResponse({"message": "Answer sentence sentiment analysis updated"})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def process_survey_sentence_sentiment(request, survey_id):
    """
    Process or reprocess all text answers in a survey for sentence-level sentiment analysis.
    
    This endpoint will update all existing text answers in a survey with 
    sentence-level sentiment analysis.
    """
    from .models import Survey, Answer
    from django.db.models import Q
    
    # Get the survey or return 404
    survey = get_object_or_404(Survey, id=survey_id)
    
    # Check if the user has permission to access this survey
    if not request.user.is_staff and survey.created_by != request.user:
        return DRFResponse({"error": "You don't have permission to access this survey"}, status=403)
    
    # Get all text answers for this survey
    answers = Answer.objects.filter(
        response__survey=survey,
        text_answer__isnull=False
    ).exclude(text_answer='')
    
    total_answers = answers.count()
    processed_answers = 0
    errors = []
    
    # Process each answer
    for answer in answers:
        try:
            # Get the language from the response
            language = answer.response.language
            
            # Analyze text at sentence level for sentiment
            from .utils import analyze_sentences, process_sentence
            sentence_data = analyze_sentences(answer.text_answer, language)
            answer.sentence_sentiments = sentence_data
            
            # For already processed answers, update sentence_sentiments and ResponseWord objects
            if answer.processed:
                # Save the sentence sentiment data
                answer.save(update_fields=['sentence_sentiments'])
                
                # Update sentence information for existing ResponseWord objects
                from .models import ResponseWord
                
                # Initialize mapping of words to sentences
                words_to_sentences = {}
                
                # Process each sentence to extract words and map them to sentences
                for idx, sentence_info in enumerate(sentence_data):
                    sentence_text = sentence_info['text']
                    sentence_idx = sentence_info['index']
                    
                    # Extract words from this sentence
                    sentence_words = process_sentence(sentence_text, language)
                    
                    # Map each word to its source sentence
                    for word in sentence_words:
                        words_to_sentences[word] = {
                            'text': sentence_text,
                            'index': sentence_idx
                        }
                
                # Get all response words for this answer
                response_words = ResponseWord.objects.filter(answer=answer)
                
                # Update sentence information for each word
                for word in response_words:
                    # Check if this word exists in our mapping
                    if word.word in words_to_sentences:
                        sentence_data = words_to_sentences[word.word]
                        word.sentence_text = sentence_data['text']
                        word.sentence_index = sentence_data['index']
                        word.save(update_fields=['sentence_text', 'sentence_index'])
            else:
                # For unprocessed answers, process them fully
                answer.process_text_answer()
                
            processed_answers += 1
        except Exception as e:
            errors.append(f"Error processing answer {answer.id}: {str(e)}")
    
    # Return summary of processing
    result = {
        "total_answers": total_answers,
        "processed_answers": processed_answers,
        "errors": errors
    }
    
    return DRFResponse(result)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def check_survey_has_responses(request, survey_id):
    """
    Check if a survey has any responses with answers.
    Returns a dictionary with information about questions that have answers.
    This is used by the frontend to warn users when deleting questions with answers.
    """
    # Get the survey or return 404
    survey = get_object_or_404(Survey, id=survey_id)
    
    # Check if the user has permission to access this survey
    if not request.user.is_staff and survey.created_by != request.user:
        return DRFResponse({"error": "You don't have permission to access this survey"}, status=403)
    
    # Get all questions for this survey that have answers
    from django.db.models import Count
    questions_with_answers = Question.objects.filter(
        survey=survey,
        answers__isnull=False
    ).annotate(
        answer_count=Count('answers')
    ).values('id', 'questions', 'answer_count')
    
    # Check if the survey has any responses
    response_count = Response.objects.filter(survey=survey).count()
    
    # Return the response with information
    return DRFResponse({
        'has_responses': response_count > 0,
        'response_count': response_count,
        'questions_with_answers': list(questions_with_answers),
        'can_delete_safely': len(questions_with_answers) == 0,
        'message': "This survey has responses with answers. Deleting questions will preserve the answer data but break the connection to the question."
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def debug_survey_questions(request, survey_id):
    """
    Debug endpoint to examine questions in a survey, showing their IDs and relationships.
    """
    # Get the survey or return 404
    survey = get_object_or_404(Survey, id=survey_id)
    
    # Check if the user has permission to access this survey
    if not request.user.is_staff and survey.created_by != request.user:
        return DRFResponse({"error": "You don't have permission to access this survey"}, status=403)
    
    # Get all questions for this survey
    questions = Question.objects.filter(survey=survey).order_by('order')
    
    # Get data about each question
    question_data = []
    for q in questions:
        # Count answers for this question
        answer_count = Answer.objects.filter(question=q).count()
        
        # Get basic info about this question
        q_info = {
            'id': q.id,
            'type': q.type,
            'order': q.order,
            'is_required': q.is_required,
            'language': q.language,
            'answer_count': answer_count,
            'created_at': q.created_at,
            'updated_at': q.updated_at
        }
        
        # Add the first language question text (for easier debugging)
        if q.questions:
            first_lang = next(iter(q.questions), None)
            if first_lang:
                q_info['first_question_text'] = q.questions.get(first_lang, '')
                q_info['language_code'] = first_lang
        
        question_data.append(q_info)
    
    # Return the response with question information
    return DRFResponse({
        'survey_id': survey.id,
        'survey_title': survey.title,
        'question_count': len(question_data),
        'questions': question_data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def test_export_endpoint(request, survey_id):
    """
    Simple test endpoint to verify API connectivity.
    """
    print("\n\n==== TEST EXPORT ENDPOINT REACHED ====")
    print(f"Survey ID: {survey_id}")
    print(f"User: {request.user}")
    print(f"Headers: {request.headers}")
    
    # Return a simple response
    return DRFResponse({
        "status": "success",
        "message": "Test export endpoint reached successfully",
        "survey_id": survey_id,
        "user": request.user.username
    })


# Keep the existing export_responses method as-is for backward compatibility, but add a new dedicated view

# Add this new function at the end of the file, after all classes and other functions

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_survey_responses_excel(request, survey_id):
    """
    Dedicated endpoint for exporting survey responses to Excel,
    with colored cells based on NPS scores.
    """
    try:
        print(f"\n==== DIRECT EXCEL EXPORT DEBUG ====")
        print(f"Request method: {request.method}")
        import pandas as pd
        import io
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill
        
        print(f"Survey ID: {survey_id}")
        print(f"User: {request.user}")
        
        # Get survey object
        survey = get_object_or_404(Survey, pk=survey_id)
        
        # Check permissions
        if not request.user.is_staff and survey.created_by != request.user:
            return HttpResponse(status=403, content="Permission denied")
        
        # Get all responses for this survey
        responses = Response.objects.filter(survey=survey).prefetch_related(
            'answers__question'
        ).order_by('-created_at')
        
        print(f"Found {responses.count()} responses to export")
        
        if not responses.exists():
            print("No responses found - returning 404")
            return HttpResponse(status=404, content="No responses found")
        
        # Prepare data for Excel export
        data = []
        
        # First, get all unique questions to use as columns
        questions = Question.objects.filter(survey=survey).order_by('order')
        print(f"Found {questions.count()} questions")
        
        # Create headers (REMOVING Response ID, Session ID as requested)
        headers = ['Date', 'Language']
        question_headers = []
        
        for question in questions:
            # Use question text in the primary survey language or the first available language
            for lang in survey.languages:
                if lang in question.questions:
                    question_headers.append(question.questions[lang])
                    break
            else:
                # Fallback if no matching language found
                question_headers.append(f"Question {question.id}")
        
        headers.extend(question_headers)
        
        # Add data for each response
        for response in responses:
            row = [
                response.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                response.language,
            ]
            
            # Create a dictionary of question_id -> answer for easy lookup
            answer_dict = {}
            for answer in response.answers.all():
                if answer.question_id:
                    answer_dict[answer.question_id] = answer
            
            # Add answers in the order of questions
            for question in questions:
                if question.id in answer_dict:
                    answer = answer_dict[question.id]
                    if answer.nps_rating is not None:
                        row.append(answer.nps_rating)
                    elif answer.text_answer:
                        row.append(answer.text_answer)
                    else:
                        row.append('')
                else:
                    row.append('')
            
            data.append(row)
        
        # Create DataFrame and Excel file
        df = pd.DataFrame(data, columns=headers)
        
        # Create a buffer for the Excel file
        buffer = io.BytesIO()
        
        # Create Excel writer and save data to it
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Responses')
            
            # Access the worksheet to apply formatting
            worksheet = writer.sheets['Responses']
            
            # Auto-adjust columns width
            for i, col in enumerate(df.columns):
                # Find the maximum length of the column
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(str(col))
                ) + 2  # Add a little extra space
                
                # Limit column width to avoid extremely wide columns
                column_width = min(max_length, 50)
                
                # Set the column width
                worksheet.column_dimensions[get_column_letter(i+1)].width = column_width
            
            # Define fill patterns for NPS scores
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Promoter (9-10)
            yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid') # Passive (7-8)
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')    # Detractor (0-6)
            
            # Identify NPS question columns
            nps_question_indices = []
            for i, question in enumerate(questions):
                if question.type == 'nps':
                    # Add 2 because Excel is 1-indexed and we have 2 metadata columns (Date, Language)
                    nps_question_indices.append(i + 2 + 1)  # +1 for Excel's 1-indexing
            
            # Color cells based on NPS scores
            for row_idx, row in enumerate(data, start=2):  # Start from 2 to skip header
                for nps_col_idx in nps_question_indices:
                    cell = worksheet.cell(row=row_idx, column=nps_col_idx)
                    try:
                        nps_value = int(cell.value) if cell.value not in (None, '') else None
                        if nps_value is not None:
                            if nps_value >= 9:
                                cell.fill = green_fill
                            elif nps_value >= 7:
                                cell.fill = yellow_fill
                            else:
                                cell.fill = red_fill
                    except (ValueError, TypeError):
                        pass  # Skip if not a valid number
        
        # Set up the response with the file
        buffer.seek(0)
        
        print(f"Excel file created, buffer size: {len(buffer.getvalue())}")
        
        # Create a plain Django HttpResponse
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="survey-responses-{survey_id}.xlsx"'
        
        print(f"Returning Excel file with headers: {dict(response.headers)}")
        return response
        
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        print(f"Error exporting responses: {str(e)}")
        print(f"Error type: {type(e).__name__}")
        print(f"Error traceback: {error_traceback}")
        logger.error(f"Error exporting responses: {str(e)}", exc_info=True)
        return HttpResponse(f"Error: {str(e)}", status=500)

# Add a new function for exporting clusters and their texts
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_survey_clusters_excel(request, survey_id):
    """
    Export clusters and their associated texts from a survey to Excel.
    Groups responses by clusters with color coding based on sentiment.
    """
    try:
        print(f"\n==== CLUSTER EXPORT DEBUG ====")
        import pandas as pd
        import io
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill
        from django.db.models import Count, Avg
        
        # Get survey object
        survey = get_object_or_404(Survey, pk=survey_id)
        
        # Check permissions
        if not request.user.is_staff and survey.created_by != request.user:
            return HttpResponse(status=403, content="Permission denied")
            
        # Get all responses filtered by survey
        responses = Response.objects.filter(survey=survey)
        
        # Get all custom clusters used in this survey's response words
        custom_clusters = CustomWordCluster.objects.filter(
            words__response__survey=survey
        ).distinct()
        
        print(f"Found {custom_clusters.count()} clusters for survey {survey_id}")
        
        if not custom_clusters.exists():
            return HttpResponse(status=404, content="No clusters found for this survey")
            
        # Create an Excel workbook with openpyxl
        buffer = io.BytesIO()
        
        # Define color fills for sentiment
        positive_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Positive
        neutral_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')   # Neutral
        negative_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Negative
        
        # Create Excel writer
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Create one sheet with all clusters and their texts
            data = []
            headers = ['Cluster', 'Text', 'Sentiment Score', 'NPS Score', 'Language']
            
            # For each cluster, collect data
            for cluster in custom_clusters:
                # Get response words for this cluster
                response_words = ResponseWord.objects.filter(
                    custom_clusters=cluster,
                    response__survey=survey
                )
                
                if not response_words.exists():
                    continue
                
                # Get unique sentences from these words
                sentences = []
                sentence_data = []
                
                # Track unique sentences to avoid duplicates
                seen_sentences = set()
                
                for word in response_words:
                    if word.sentence_index is None or not word.answer:
                        continue
                        
                    # Get the sentence if available
                    sentence_text = None
                    sentiment_score = 0
                    
                    if word.answer.sentence_sentiments:
                        for sent in word.answer.sentence_sentiments:
                            if sent.get('index') == word.sentence_index:
                                sentence_text = sent.get('text')
                                sentiment_score = sent.get('sentiment', 0)
                                break
                    
                    # If we found a valid sentence and haven't seen it before
                    if sentence_text and (word.answer_id, word.sentence_index) not in seen_sentences:
                        seen_sentences.add((word.answer_id, word.sentence_index))
                        
                        # Get NPS score if available
                        nps_score = None
                        nps_answers = Answer.objects.filter(
                            response=word.response,
                            question__type='nps',
                            nps_rating__isnull=False
                        ).first()
                        
                        if nps_answers:
                            nps_score = nps_answers.nps_rating
                        
                        # Add to our data array
                        data.append([
                            cluster.name,
                            sentence_text,
                            sentiment_score,
                            nps_score,
                            word.response.language
                        ])
            
            # Create the dataframe and add to Excel
            if data:
                df = pd.DataFrame(data, columns=headers)
                df.to_excel(writer, index=False, sheet_name='Clusters')
                
                # Access the worksheet to format cells
                worksheet = writer.sheets['Clusters']
                
                # Auto-adjust column widths
                for i, col in enumerate(df.columns):
                    max_length = max(
                        df[col].astype(str).map(len).max(),
                        len(str(col))
                    ) + 2
                    column_width = min(max_length, 50)
                    worksheet.column_dimensions[get_column_letter(i+1)].width = column_width
                
                # Color cells based on sentiment
                for row_idx, row_data in enumerate(data, start=2):  # Start from 2 to skip header
                    sentiment_score = row_data[2]  # Index 2 is the sentiment score
                    cell = worksheet.cell(row=row_idx, column=3)  # Column 3 is the sentiment column
                    
                    if sentiment_score > 0.3:
                        cell.fill = positive_fill
                    elif sentiment_score < -0.3:
                        cell.fill = negative_fill
                    else:
                        cell.fill = neutral_fill
                    
                    # Also color the NPS column
                    nps_score = row_data[3]  # Index 3 is the NPS score
                    if nps_score is not None:
                        nps_cell = worksheet.cell(row=row_idx, column=4)  # Column 4 is the NPS column
                        if nps_score >= 9:
                            nps_cell.fill = positive_fill
                        elif nps_score >= 7:
                            nps_cell.fill = neutral_fill
                        else:
                            nps_cell.fill = negative_fill
            else:
                # Create an empty sheet if no data
                pd.DataFrame([], columns=headers).to_excel(writer, index=False, sheet_name='Clusters')
                worksheet = writer.sheets['Clusters']
                for i, col in enumerate(headers):
                    worksheet.column_dimensions[get_column_letter(i+1)].width = len(col) + 5
            
            # Create a second sheet with cluster summary statistics
            summary_data = []
            summary_headers = ['Cluster', 'Frequency', 'Avg Sentiment', 'Avg NPS', 'Keywords']
            
            for cluster in custom_clusters:
                # Get words for this cluster
                response_words = ResponseWord.objects.filter(
                    custom_clusters=cluster,
                    response__survey=survey
                )
                
                if not response_words.exists():
                    continue
                
                # Count unique sentence occurrences
                unique_sentences = set()
                for word in response_words:
                    if word.sentence_index is not None:
                        unique_sentences.add((word.response_id, word.answer_id, word.sentence_index))
                
                # Frequency is the number of unique sentences
                frequency = len(unique_sentences)
                
                # Calculate average sentiment
                avg_sentiment = response_words.aggregate(Avg('sentiment_score'))['sentiment_score__avg'] or 0
                
                # Get average NPS score
                avg_nps = None
                nps_answers = Answer.objects.filter(
                    response__in=response_words.values('response').distinct(),
                    question__type='nps',
                    nps_rating__isnull=False
                )
                
                if nps_answers.exists():
                    nps_scores = list(nps_answers.values_list('nps_rating', flat=True))
                    avg_nps = sum(nps_scores) / len(nps_scores) if nps_scores else None
                
                # Get keywords string
                keywords = ', '.join(cluster.keywords[:5]) if cluster.keywords else ''
                
                # Add to summary data
                summary_data.append([
                    cluster.name,
                    frequency,
                    round(avg_sentiment, 2),
                    round(avg_nps, 1) if avg_nps is not None else None,
                    keywords
                ])
            
            # Create summary sheet
            if summary_data:
                summary_df = pd.DataFrame(summary_data, columns=summary_headers)
                summary_df.to_excel(writer, index=False, sheet_name='Summary')
                
                # Format the summary sheet
                summary_sheet = writer.sheets['Summary']
                for i, col in enumerate(summary_headers):
                    max_length = max(
                        summary_df[col].astype(str).map(len).max(),
                        len(str(col))
                    ) + 2
                    column_width = min(max_length, 50)
                    summary_sheet.column_dimensions[get_column_letter(i+1)].width = column_width
                
                # Color sentiment and NPS cells
                for row_idx, row_data in enumerate(summary_data, start=2):
                    # Color sentiment cell
                    sentiment_score = row_data[2]  # Index 2 is the avg sentiment score
                    sentiment_cell = summary_sheet.cell(row=row_idx, column=3)
                    
                    if sentiment_score > 0.3:
                        sentiment_cell.fill = positive_fill
                    elif sentiment_score < -0.3:
                        sentiment_cell.fill = negative_fill
                    else:
                        sentiment_cell.fill = neutral_fill
                    
                    # Color NPS cell
                    nps_score = row_data[3]  # Index 3 is the avg NPS score
                    if nps_score is not None:
                        nps_cell = summary_sheet.cell(row=row_idx, column=4)
                        if nps_score >= 9:
                            nps_cell.fill = positive_fill
                        elif nps_score >= 7:
                            nps_cell.fill = neutral_fill
                        else:
                            nps_cell.fill = negative_fill
        
        # Set up the response with the file
        buffer.seek(0)
        
        print(f"Cluster Excel file created, buffer size: {len(buffer.getvalue())}")
        
        # Create a plain Django HttpResponse
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="survey-clusters-{survey_id}.xlsx"'
        
        return response
        
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        print(f"Error exporting clusters: {str(e)}")
        print(f"Error traceback: {error_traceback}")
        logger.error(f"Error exporting clusters: {str(e)}", exc_info=True)
        return HttpResponse(f"Error: {str(e)}", status=500)


